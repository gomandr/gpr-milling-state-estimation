import gpytorch
import numpy as np
import openpyxl
import time
import torch

from error_metrics import error_metrics
from matplotlib import pyplot as plt
from models import Matern_Kernel, SE_Kernel, RQ_Kernel, PERxRQaddSE_Kernel
from preprocessing import preprocessing_part1, preprocessing_part2, cross_validation_split

##############################################################################
### Parameters

### Dataset -> Needs small manual changes in create_dataset.py for selecting with/without theta
dataset_name = 'shift + random'
#dataset_name = 'rotated_z'

#path_to_data = ['Prototype.npy'] 
path_to_data = ['DATASETS/Shift_x/*.npy', 'DATASETS/Random_action/*.npy']  
#path_to_data = ['DATASETS/Rotated_z/*.npy']

#path_to_data = ['DATASETS/Shift_y/*.npy']                  # Requires theta in 'create_dataset.py'
#path_to_data = ['DATASETS/Shift_y_and_random/*.npy']       # Requires theta in 'create_dataset.py'

### Preprocessing parameters
threshold_corr = 1                      # Spearman's rank correlation coefficient threshold  
preproc_ps0='delete'               # Action on 0 values of ps input: 'delete', 'interpolate' or 'None'
data_augmentation = True                # True: adds previous timesteps as additional features

### Training parameters
training_iter = 1000                    # Maximal number of optimisation iterations
learning_rate = 0.01                    # Learning rate of optimiser (for both SGD and ADAM)
momentum = 0.9                          # Momentum  of optimiser (for SGD)
cross_validation = True                 # True: activates K-fold cross-validation
folds = 3                               # Number of folds for K-fold cv
y_scrambling = False                    # True: activates y-scrambling 

### Kernel - covariance function parameters 
kernel = 'RQ'                           # RQ, SE, Matern, ...
nu = 0.5                                # Smoothness (for Matérn kernel only)
dims_kernel = 'all'                     # Lengthscale per dimension: one overall or one per feature 
hypers = {                              # Non-default initialisation of hyperparameters
    'likelihood.noise_covar.noise': torch.tensor(0.1)
    }

### Preload of model, training/retraining parameters
preload = False                         # Use preloaded model (skip training and goes directly to prediction)
train = True                            # Train model or not
preload_model_name = 'Models/all_matern_all_2dot5_lr0dot01_thresh0dot9_isRandomUseless.pth'

### Print and plot options
print_iterations = True                 # Training iteration info
print_error = True                      # Final error on prediction
show_plot = False                       # Plots of prediction for each input feature

### Saving model/scoring metrics
save = False                            # Save the model at the end 
save_name = 'Models/all_RQ_all__lr0dot01_thresh1_isRandomUseless.pth'
save_info = True                        # Save scoring metrics in Excel file
wbkName = 'results_cv.xlsx' 

##############################################################################
### Preprocessing

train_val_ds, test_ds = preprocessing_part1(path_to_data=path_to_data, \
                        data_augmentation=data_augmentation, zeros_action=preproc_ps0)

if cross_validation == True: # Computes folds indexes sets
    idx_split = cross_validation_split(len(train_val_ds), folds=folds)
else:
    folds = 1

##############################################################################
### Cross-validation loop

mase_train_folds = []
mase_valid_folds = []
naive_train_folds = []
naive_valid_folds = []
mape_train_folds = []
mape_valid_folds = []
mae_train_folds = []
mae_valid_folds = []
            
for i in range(folds):
    print('### Fold ', str(i+1), '/', str(folds))
    if cross_validation == True: # Puts the right set in train an valid set     
        indices = list(range(folds))
        ind_train = indices[:i] + indices[i+1:]
        
        idx_val = idx_split[i]
        idx_train = [idx_split[i] for i in ind_train]
        idx_train = sum(idx_train, [])
        
        train_ds = train_val_ds.loc[idx_train]
        val_ds = train_val_ds.loc[idx_val]
        
    Xn_train, Yn_train, Xn_val, Yn_val, Xn_test, Yn_test, mean, var, headers \
    = preprocessing_part2(train_ds.copy(), val_ds.copy(), test_ds.copy(), threshold_corr=threshold_corr)
    

    # Scaling metrics and destandardisation of targets for computing error
    std_x = np.sqrt(var[:-1])
    std_y = np.sqrt(var[-1])
    mean_x = mean[:-1]
    mean_y = mean[-1]
    
    train_y =  Yn_train*std_y + mean_y
    valid_x = Xn_val*std_x + mean_x
    valid_y = Yn_val*std_y + mean_y
    
  
    if y_scrambling == True: 
        np.random.shuffle(Yn_train)
    
    # Convert to torch
    train_xn = torch.from_numpy(Xn_train)
    valid_xn = torch.from_numpy(Xn_val)
    train_yn = torch.from_numpy(Yn_train)
    valid_yn = torch.from_numpy(Yn_val)

    
    ##############################################################################
    ### Model
    
    # Initialize likelihood and model
    print('Creating new model...')
    likelihood = gpytorch.likelihoods.GaussianLikelihood() 
    if kernel == 'Matern': 
        model = Matern_Kernel(train_xn, train_yn, likelihood, nu=nu, dims=dims_kernel)
    elif kernel == 'SE':
        model = SE_Kernel(train_xn, train_yn, likelihood, dims=dims_kernel)
    elif kernel == 'RQ':
        model = RQ_Kernel(train_xn, train_yn, likelihood, dims=dims_kernel)
    elif kernel == 'PERxRQaddSE':
        model =PERxRQaddSE_Kernel(train_xn, train_yn, likelihood)
         
    if preload == True:
        print('Loading saved parameters...')
        state_dict = torch.load(preload_model_name)
        model.load_state_dict(state_dict)  
    else:
        model.initialize(**hypers)
        
    ###########################################################################
    ### Optimizer and training loop

    loss_history = []
    t0 = time.time()
    if  train == True:
        print('Training the model...')
        model.train()
        model = model.double()
        likelihood.train()
        
        # Define optimizer
        ''' # SGD 
        optimizer = torch.optim.SGD(model.parameters(), lr=learning_rate, momentum=momentum)
        '''
        # ADAM
        optimizer = torch.optim.Adam([
            {'params': model.parameters()},  # Includes GaussianLikelihood parameters
        ], lr=learning_rate)
        
        
        # Loss - Marginal Log Likelihood
        mll = gpytorch.mlls.ExactMarginalLogLikelihood(likelihood, model)
        
        best_valid_loss = np.inf
        it_not_improved = 0
        
        # Training loop
        for i in range(training_iter):
            # Zero gradients from previous iteration
            optimizer.zero_grad()
            output = model(train_xn)
            
            # Calc loss and backprop gradients
            loss = -mll(output, train_yn)
            loss.backward()
            
            
            if (i+1)%1 == 0: # Decides how often to validate and save model
                model.eval()
                likelihood.eval()
                model = model.double()
    
                # Make predictions by feeding model through likelihood
                with torch.no_grad(), gpytorch.settings.fast_pred_var():
                    output = model(valid_xn)
                    loss_valid = -mll(output, valid_yn) 
                    if print_iterations == True:
                        print('Iter %d/%d - Train loss: %.2e - Validation loss: %.2e' % ( #' - Lengthscales: %.3f' % (
                        i + 1, training_iter, loss.item(), loss_valid.item() # , model.covar_module.base_kernel.lengthscale[0,0].item()
                        ))
                    loss_history.append([loss.item(), loss_valid.item()])
                
                # Record if better results
                if loss_valid < best_valid_loss:
                    best_iter = i
                    best_valid_loss = loss_valid
                    best_train_loss = loss
                    best_model = model
                    best_likelihood = likelihood
                    it_not_improved = 0
                    
                    if save == True:  
                        print('Saving the trained model...')
                        torch.save(model.state_dict(), save_name)
                else:
                    it_not_improved += 1
                    
                    if it_not_improved == 10:
                        break
                 
                model.train()
                model = model.double()
                likelihood.train()
                    
            optimizer.step()
            
        t1 = time.time()
        training_time = t1-t0
        print('Training time: %.1f' % (training_time))
        print('Best model: iteration: %d - training loss: %.2e - validation loss: %.2e' % (
                            best_iter+1, best_train_loss, best_valid_loss
                            ))
    else: # No training 
        best_model = model
        best_likelihood = likelihood
        best_iter = -1
        training_time = -1
        best_train_loss = -1
        best_valid_loss = -1
        
    ##############################################################################
    ### Evaluation
    
    # Get into evaluation (predictive posterior) mode
    print('Predicting...')
    best_model.eval()
    best_likelihood.eval()
    best_model = best_model.double()
    
    # Make predictions by feeding model through likelihood
    with torch.no_grad(), gpytorch.settings.fast_pred_var():
        # Predictions for training data
        observed_pred_train = best_likelihood(best_model(train_xn))
        pred_train_yn = observed_pred_train.mean.numpy()
        
        # Predictions for validation data
        observed_pred_valid = best_likelihood(best_model(valid_xn))  
        pred_valid_yn = observed_pred_valid.mean.numpy()
        
        # Return to original scale 
        pred_valid_y = pred_valid_yn * std_y + mean_y
        pred_train_y = pred_train_yn * std_y + mean_y
        
        # Compute errors (mase, naive, mape) with cross-validation metrics (mean, std, relative std)       
        mape_train, mape_valid, mae_train, mae_valid, mase_train,\
        mase_valid, naive_train, naive_valid  = error_metrics(pred_train_y, \
                                                train_y, pred_valid_y, valid_y, \
                                                std_y, print_error=True, first=True, last=True)

        mase_train_folds = np.append(mase_train_folds, mase_train)
        mase_valid_folds = np.append(mase_valid_folds, mase_valid)
        naive_train_folds = np.append(naive_train_folds, naive_train)
        naive_valid_folds = np.append(naive_valid_folds, naive_valid)
        mape_train_folds = np.append(mape_train_folds, mape_train)
        mape_valid_folds = np.append(mape_valid_folds, mape_valid)
        mae_train_folds = np.append(mae_train_folds, mae_train)
        mae_valid_folds = np.append(mae_valid_folds, mae_valid)
        
        mase_cv = [np.mean(mase_train_folds), np.mean(mase_valid_folds), \
                  np.std(mase_train_folds), np.std(mase_valid_folds)]
        naive_cv = [np.mean(naive_train_folds), np.mean(naive_valid_folds), \
                  np.std(naive_train_folds), np.std(naive_valid_folds)]
        mape_cv = [np.mean(mape_train_folds), np.mean(mape_valid_folds), \
                  np.std(mape_train_folds), np.std(mape_valid_folds)]        
        mae_cv = [np.mean(mae_train_folds), np.mean(mae_valid_folds), \
              np.std(mae_train_folds), np.std(mae_valid_folds)]
             
        print('MASE: \n Training: mean = %.2e, std = %.2e, rsd = %.2f %% \
              \n Validation: mean = %.2e, std = %.2e, rsd = %.2f %%' % (
              mase_cv[0], mase_cv[2], mase_cv[2]*100/mase_cv[0], \
                  mase_cv[1], mase_cv[3], mase_cv[3]*100/mase_cv[1]
        ))
        
        print('Naive: \n Training: mean = %.2e, std = %.2e, rsd = %.2f %% \
              \n Validation: mean = %.2e, std = %.2e, rsd = %.2f %%' % (
              naive_cv[0], naive_cv[2], naive_cv[2]*100/naive_cv[0], \
                  naive_cv[1], naive_cv[3], naive_cv[3]*100/naive_cv[1]
        ))
        
        print('MAPE: \n Training: mean = %.2e, std = %.2e, rsd = %.2f %% \
              \n Validation: mean = %.2e, std = %.2e, rsd = %.2f %%' % (
              mape_cv[0], mape_cv[2], mape_cv[2]*100/mape_cv[0], \
                  mape_cv[1], mape_cv[3], mape_cv[3]*100/mape_cv[1]
        ))
    
        print('MAE: \n Training: mean = %.2e, std = %.2e, rsd = %.2f %% \
              \n Validation: mean = %.2e, std = %.2e, rsd = %.2f %%' % (
              mae_cv[0], mae_cv[2], mae_cv[2]*100/mae_cv[0], \
                  mae_cv[1], mae_cv[3], mae_cv[3]*100/mae_cv[1]
        ))      
    
    
        # Compute confidence region (+/- 2*std)
        lower, upper = observed_pred_valid.confidence_region()
        lower_err =  (pred_valid_yn - lower.numpy())* std_y
        upper_err = (upper.numpy() - pred_valid_yn) * std_y
        
##############################################################################
### Save info
if save_info==True:
    
    if kernel == 'Matern':
        model_name = str(kernel) + ', nu=' + str(nu) + ', ls=' + dims_kernel
    else:
        model_name = str(kernel) + ', ls=' + dims_kernel
         
    wbk = openpyxl.load_workbook(wbkName)
    ws = wbk.worksheets[0]
    for cell in ws["A"]:
        if cell.value is None:
            idx = cell.row
            break
    else:
        idx = cell.row + 1 
    
    values_to_write = [dataset_name, model_name, best_iter+1, learning_rate, \
                        training_time, threshold_corr, preproc_ps0, \
                        mae_cv[0], mae_cv[2], mae_cv[2]*100/mae_cv[0], \
                        mae_cv[1], mae_cv[3], mae_cv[3]*100/mae_cv[1], \
                        mape_cv[0], mape_cv[2], mape_cv[2]*100/mape_cv[0], \
                        mape_cv[1], mape_cv[3], mape_cv[3]*100/mape_cv[1], \
                        mase_cv[0], mase_cv[2], mase_cv[2]*100/mase_cv[0], \
                        mase_cv[1], mase_cv[3], mase_cv[3]*100/mase_cv[1], \
                        naive_cv[0], naive_cv[2], naive_cv[2]*100/naive_cv[0], \
                        naive_cv[1], naive_cv[3], naive_cv[3]*100/naive_cv[1], \
                        (best_train_loss/len(Yn_train)).item(), \
                        (best_valid_loss/len(Yn_val)).item()]
        
    for i, value in enumerate(values_to_write):
        ws.cell(row=idx,column=i+1).value = value
        
    wbk.save(wbkName)
    wbk.close

##############################################################################

if show_plot == True:    
    print('Plotting ...')
    
    # Loss evolution
    fig = plt.figure()
    ax0 = fig.add_subplot(121, title="Losses (-MLL)")
    for i, losses in enumerate(loss_history):
        ax0.plot(i, np.exp(losses[0])/len(Yn_train), 'bo', markersize=3, label='training')
        ax0.plot(i, np.exp(losses[1])/len(Yn_val), 'rx', markersize=3, label='validation')
        if i == 0:
            ax0.legend()
            ax0.set_xlabel('# Iteration')
            ax0.set_ylabel('Loss')
    plt.grid()
        #fig.savefig( os.path.join(save_dir, name,'train.jpg'))

    # 2D-plot prediction along chosen dimensions 
    headers_x = headers[:-1]
    indices = list(range(len(valid_x)))
    np.random.shuffle(indices)
    ratio = 0.1
    split = int(np.floor(len(valid_x) * ratio))
    for i in [0,1,2,-4,]:                     #range(2):#valid_x.shape[1])
        
    
        f, ax = plt.subplots(1, 1, figsize=(8, 6))
        #plt.title('Prediction of quality deviation on validation set')
        ax.errorbar(valid_x[:split,i], pred_valid_y[:split], \
                    yerr=[lower_err[:split], upper_err[:split]], fmt='o', ms=2, elinewidth=1)
        ax.plot(valid_x[:split,i], valid_y[:split], 'rx', ms=2)
        ax.legend(['True value', 'Predicted mean with confidence'])
        ax.set_xlabel(str(headers_x[i]))
        ax.set_ylabel('quality deviation [mm]')
        ax.grid(b=True)
        
    # 3D-plot 
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    ax.scatter(valid_x[:,0], valid_x[:,2], valid_y)
    ax.scatter(valid_x[:,0], valid_x[:,2], pred_valid_y)